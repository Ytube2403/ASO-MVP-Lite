import os
import re
import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_OUTPUT_DIR = os.path.join(PROJECT_ROOT, "apps", "ElectricGun", "Output", "062026")

def get_first(row_data, *keys):
    for key in keys:
        value = row_data.get(key)
        if value not in (None, ""):
            return value
    return None


def parse_args():
    parser = argparse.ArgumentParser(description="Consolidate ASO shortlist workbook sheets into one overview workbook.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Directory containing locale output .xlsx files.")
    parser.add_argument("--output-name", default="ElectricGun_Consolidated_Shortlists.xlsx", help="Consolidated workbook filename.")
    return parser.parse_args()


def consolidate(output_dir=DEFAULT_OUTPUT_DIR, output_name="ElectricGun_Consolidated_Shortlists.xlsx"):
    print("Consolidating Main Shortlists...")
    if not os.path.exists(output_dir):
        print(f"Error: Output directory {output_dir} does not exist.")
        return

    # Find all Excel output files
    files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx") and not f.startswith("~$") and f != output_name]

    # Sort files to keep consistent order
    files.sort()

    if not files:
        print("No output Excel files found to consolidate.")
        return

    # Create new workbook for consolidation
    consolidated_wb = Workbook()

    # Remove default sheet
    default_sheet = consolidated_wb.active
    consolidated_wb.remove(default_sheet)

    # Master Overview Sheet
    overview_ws = consolidated_wb.create_sheet(title="Master Overview")
    overview_ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    regular_font = Font(name=font_family, size=11)
    bold_font = Font(name=font_family, size=11, bold=True)

    thin_border_side = Side(border_style="thin", color="DADCE0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Overview headers
    headers = ["Locale", "Rank", "Keyword", "Volume", "Difficulty", "KEI", "RelevancyScore", "BalancedScore", "Bucket", "Reason"]
    for col_idx, header in enumerate(headers, 1):
        cell = overview_ws.cell(row=1, column=col_idx, value=header)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    overview_row = 2

    for file_name in files:
        # Extract locale from filename (e.g. electric_gun_BR-PT_Output.xlsx -> BR-PT)
        match = re.search(r'_([A-Z]{2}-[A-Z]{2})_Output\.xlsx$', file_name, re.IGNORECASE)
        if not match:
            # Fallback check
            match = re.search(r'([A-Z]{2}-[A-Z]{2})', file_name)

        locale = match.group(1).upper() if match else file_name.replace(".xlsx", "")
        file_path = os.path.join(output_dir, file_name)

        print(f"Processing locale: {locale} ({file_name})...")

        # Load file
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"  Error loading {file_name}: {e}")
            continue

        # Look for the main shortlist sheet
        shortlist_sheet_name = next((name for name in wb.sheetnames if "Shortlist" in name or "01_" in name), None)
        if not shortlist_sheet_name:
            print(f"  Warning: No shortlist sheet found in {file_name}")
            wb.close()
            continue

        ws = wb[shortlist_sheet_name]

        # We also create a separate sheet for this locale in the consolidated workbook
        locale_ws = consolidated_wb.create_sheet(title=locale)
        locale_ws.views.sheetView[0].showGridLines = True

        # Copy headers from source sheet to locale sheet
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=1, column=col_idx).value
            cell = locale_ws.cell(row=1, column=col_idx, value=val)
            cell.font = title_font
            cell.fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid") # Green header for locales
            cell.alignment = align_center
            cell.border = thin_border

        # Read rows
        row_idx = 2
        for r in range(2, ws.max_row + 1):
            # Check if keyword is empty
            kw_val = ws.cell(row=r, column=2).value
            if not kw_val:
                continue

            # 1. Copy row to the specific locale sheet
            for c in range(1, max_col + 1):
                src_cell = ws.cell(row=r, column=c)
                dest_cell = locale_ws.cell(row=row_idx, column=c, value=src_cell.value)
                dest_cell.font = regular_font
                dest_cell.border = thin_border
                if c in [1, 3, 4, 5]: # Rank, Volume, Diff, KEI
                    dest_cell.alignment = align_center
                else:
                    dest_cell.alignment = align_left

            # 2. Extract values for Master Overview
            # We map source columns by name to be safe
            row_data = {}
            for c in range(1, max_col + 1):
                header_name = str(ws.cell(row=1, column=c).value or "").strip().lower()
                row_data[header_name] = ws.cell(row=r, column=c).value

            # Storing to Overview sheet
            overview_ws.cell(row=overview_row, column=1, value=locale).alignment = align_center
            overview_ws.cell(row=overview_row, column=2, value=get_first(row_data, "rank", "#")).alignment = align_center
            overview_ws.cell(row=overview_row, column=3, value=row_data.get("keyword"))
            overview_ws.cell(row=overview_row, column=4, value=row_data.get("volume")).alignment = align_center
            overview_ws.cell(row=overview_row, column=5, value=row_data.get("difficulty")).alignment = align_center
            overview_ws.cell(row=overview_row, column=6, value=row_data.get("kei")).alignment = align_center

            # Format floats
            rel_score = row_data.get("relevancyscore")
            if isinstance(rel_score, (int, float)):
                overview_ws.cell(row=overview_row, column=7, value=round(rel_score, 4)).alignment = align_center
            else:
                overview_ws.cell(row=overview_row, column=7, value=rel_score).alignment = align_center

            bal_score = row_data.get("balancedscore")
            if isinstance(bal_score, (int, float)):
                overview_ws.cell(row=overview_row, column=8, value=round(bal_score, 4)).alignment = align_center
            else:
                overview_ws.cell(row=overview_row, column=8, value=bal_score).alignment = align_center

            overview_ws.cell(row=overview_row, column=9, value=get_first(row_data, "bucket", "group", "section")).alignment = align_center
            overview_ws.cell(row=overview_row, column=10, value=get_first(row_data, "reason", "decisionreason"))

            # Format overview row cells
            for c in range(1, 11):
                cell = overview_ws.cell(row=overview_row, column=c)
                cell.font = regular_font
                cell.border = thin_border
                if c not in [1, 2, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = align_left

            overview_row += 1
            row_idx += 1

        # Autofit columns for locale sheet
        for col in locale_ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            locale_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.close()

    # Autofit columns for overview sheet
    for col in overview_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        overview_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save consolidated workbook
    dest_path = os.path.join(output_dir, output_name)
    consolidated_wb.save(dest_path)
    print(f"Consolidated file saved successfully at: {dest_path}")

if __name__ == "__main__":
    args = parse_args()
    consolidate(args.output_dir, args.output_name)
