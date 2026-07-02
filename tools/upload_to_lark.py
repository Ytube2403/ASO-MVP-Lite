import os
import sys
import json
import urllib.request
import urllib.error
import argparse
import openpyxl

def get_tenant_access_token(app_id, app_secret):
    url = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
    headers = {"Content-Type": "application/json; charset=utf-8"}
    data = {"app_id": app_id, "app_secret": app_secret}
    
    req = urllib.request.Request(
        url, 
        data=json.dumps(data).encode("utf-8"), 
        headers=headers, 
        method="POST"
    )
    
    try:
        with urllib.request.urlopen(req) as resp:
            res_data = json.loads(resp.read().decode("utf-8"))
            if res_data.get("code") == 0:
                return res_data["tenant_access_token"]
            else:
                raise ValueError(f"Error getting token: {res_data.get('msg')}")
    except urllib.error.URLError as e:
        raise ConnectionError(f"HTTP Connection failed: {e}")

def upload_sheet_data(token, spreadsheet_token, sheet_id, start_cell, end_cell, values):
    url = f"https://open.larksuite.com/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8"
    }
    
    range_str = f"{sheet_id}!{start_cell}:{end_cell}"
    payload = {
        "valueRange": {
            "range": range_str,
            "values": values
        }
    }
    
    req = urllib.request.Request(
        url, 
        data=json.dumps(payload).encode("utf-8"), 
        headers=headers, 
        method="PUT"
    )
    
    try:
        with urllib.request.urlopen(req) as resp:
            res_data = json.loads(resp.read().decode("utf-8"))
            if res_data.get("code") == 0:
                print(f"SUCCESS: Successfully updated range {range_str}!")
            else:
                raise ValueError(f"Error writing data: {res_data.get('msg')}")
    except urllib.error.URLError as e:
        raise ConnectionError(f"HTTP Connection failed: {e}")

def main():
    parser = argparse.ArgumentParser(description="Upload ASO Keyword Shortlist to Lark Sheet")
    parser.add_argument("--excel", type=str, required=True, help="Path to ASO Output Excel")
    parser.add_argument("--app-id", type=str, required=True, help="Lark App ID")
    parser.add_argument("--app-secret", type=str, required=True, help="Lark App Secret")
    parser.add_argument("--token", type=str, required=True, help="Lark Spreadsheet Token")
    parser.add_argument("--sheet-id", type=str, required=True, help="Lark Sheet name/ID (e.g. 'US-EN' or 'Sheet3')")
    parser.add_argument("--sheet-tab", type=str, default="01_Main_Keyword_Shortlist", help="Excel sheet tab to read from")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"ERROR: Excel file not found: {args.excel}")
        sys.exit(1)
        
    print(f"Reading from Excel sheet tab '{args.sheet_tab}'...")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    if args.sheet_tab not in wb.sheetnames:
        print(f"ERROR: Excel does not contain tab '{args.sheet_tab}'. Available tabs: {wb.sheetnames}")
        sys.exit(1)
        
    ws = wb[args.sheet_tab]
    
    # Extract header + rows
    values = []
    # Determine the columns to extract: A to H
    for row in range(1, ws.max_row + 1):
        row_vals = []
        for col in range(1, 9): # A=1 to H=8
            val = ws.cell(row=row, column=col).value
            # convert none or formatting to clean string/number
            if val is None:
                row_vals.append("")
            else:
                row_vals.append(val)
        values.append(row_vals)
        
    wb.close()
    
    print(f"Extracted {len(values)} rows (including header) from Excel.")
    
    if not values:
        print("ERROR: No data to upload.")
        sys.exit(1)
        
    print("Authenticating with Lark Open Platform...")
    token = get_tenant_access_token(args.app_id, args.app_secret)
    print("Token obtained successfully.")
    
    # Calculate end cell (e.g., column H, row count)
    end_row = len(values)
    start_cell = "A1"
    end_cell = f"H{end_row}"
    
    print(f"Uploading values to Lark sheet '{args.sheet_id}' range {start_cell}:{end_cell}...")
    upload_sheet_data(token, args.token, args.sheet_id, start_cell, end_cell, values)
    print("Upload complete!")

if __name__ == "__main__":
    main()
