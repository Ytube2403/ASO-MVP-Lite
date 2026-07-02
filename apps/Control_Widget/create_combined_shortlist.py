import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import datetime

# Parse month argument
parser = argparse.ArgumentParser(description="Create combined shortlist for Control Widget")
parser.add_argument("--month", type=str, default="", help="Month in MMYYYY format (e.g. 072026)")
args = parser.parse_args()

month = args.month
if not month:
    month = datetime.datetime.now().strftime("%m%Y")

# Define paths
APP_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(APP_DIR, "Output", month)
file_paths = {
    "BR_PT": os.path.join(OUTPUT_DIR, "control_widget_BR-PT_Output.xlsx"),
    "MX_ES": os.path.join(OUTPUT_DIR, "control_widget_MX-ES_Output.xlsx"),
    "US_EN": os.path.join(OUTPUT_DIR, "control_widget_US-EN_Output.xlsx")
}

# Output file path
combined_file_path = os.path.join(OUTPUT_DIR, "control_widget_Combined_Shortlist.xlsx")

# Workbook initialization
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove active sheet

# Styling function similar to the main pipeline styling
def style_sheet(ws, title, is_combined=False):
    ws.views.sheetView[0].showGridLines = True
    if ws.max_row > 1:
        ws.freeze_panes = 'A2'

    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Style header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            col_name = ws.cell(row=1, column=col_idx).value
            
            # Numeric conversion and formatting
            if col_name in ['Volume', 'Max. Volume', 'Difficulty', 'Rank', 'MaximumReach']:
                try:
                    cell.value = int(float(cell.value))
                except:
                    pass
            elif col_name in ['BalancedScore', 'RelevancyScore', 'KEI', 'VolumeN', 'DifficultyN', 'KEIN', 'CurrentRankN', 'OpportunityRankGap', 'ExpansionValue', 'Traffic Stability']:
                try:
                    cell.value = round(float(cell.value), 4)
                    if col_name == 'Traffic Stability':
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '0.0000'
                except:
                    pass

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Hide Traffic Stability and Stability Class if present
        col_name = ws.cell(row=1, column=col[0].column).value
        if col_name in ['Traffic Stability', 'Stability Class']:
            ws.column_dimensions[col_letter].hidden = True

# Read sheets and build lists
combined_data = []

# First sheet: Combined List
for locale, path in file_paths.items():
    if not os.path.exists(path):
        print(f"Warning: File {path} does not exist. Skipping.")
        continue
    
    # Load sheet
    df_locale = pd.read_excel(path, sheet_name="01_Main_Keyword_Shortlist")
    
    # Insert Locale as the first column
    df_locale.insert(0, 'Locale', locale)
    combined_data.append(df_locale)

if combined_data:
    df_all_combined = pd.concat(combined_data, ignore_index=True)
    
    # Create combined sheet in Excel
    ws_combined = wb.create_sheet(title="All_Locales_Combined")
    
    # Write headers
    headers = list(df_all_combined.columns)
    for col_idx, col_name in enumerate(headers, 1):
        ws_combined.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, row in df_all_combined.iterrows():
        for col_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            # Replace nan with None
            if pd.isna(val):
                val = None
            ws_combined.cell(row=row_idx + 2, column=col_idx, value=val)
            
    style_sheet(ws_combined, "All_Locales_Combined", is_combined=True)

# Individual sheets
for locale, path in file_paths.items():
    if not os.path.exists(path):
        continue
    
    df_locale = pd.read_excel(path, sheet_name="01_Main_Keyword_Shortlist")
    ws_locale = wb.create_sheet(title=f"{locale}_Main_List")
    
    headers = list(df_locale.columns)
    for col_idx, col_name in enumerate(headers, 1):
        ws_locale.cell(row=1, column=col_idx, value=col_name)
        
    for row_idx, row in df_locale.iterrows():
        for col_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            ws_locale.cell(row=row_idx + 2, column=col_idx, value=val)
            
    style_sheet(ws_locale, f"{locale}_Main_List")

# Save combined workbook
wb.save(combined_file_path)
print(f"Combined Excel report saved successfully to: {combined_file_path}")
